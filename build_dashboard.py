#!/usr/bin/env python3
"""
Genera dashboard.html a partir de un Excel de seguimiento de eventos.

Uso:
    python3 build_dashboard.py ruta/al/excel.xlsx [ruta/salida.html]

Si no se da ruta de salida, escribe "dashboard.html" en el directorio actual.
Requiere que "template.html" (el HTML del dashboard) esté en el mismo directorio
que este script, o ajusta TEMPLATE_PATH abajo.

Requiere: pip install openpyxl --break-system-packages
"""
import sys
import json
import datetime
import openpyxl

TEMPLATE_PATH = "template.html"

# Nombres EXACTOS de columnas que espera el dashboard (deben coincidir con tu Excel)
COLUMNS = [
    "Revisión", "Acción Actual", "EVENTO+PONENTE", "Número de evento",
    "Estatus del Evento", "AT", "Tipo de Evento", "Franquicia",
    "Nombre del evento", "Nombre Speaker", "Estatus del Speaker", "ID Vendor",
    "Régimen fiscal", "Fecha del evento", "Year", "Month",
    "Estatus de Contrato", "Número ICD contrato", "Request Navigator",
    "Attendee ID", "Solicitud de excepción", "Logística del speaker",
    "Fecha solicitud del evento", "Fecha de firma de contrato",
    "Nombre del Dueño del evento", "Dirección correo elec. ponente/consultor",
    "Fecha de solicitud de pago en sistema", "Producto/Marca", "CECO", "OI",
    "ACCOUNT", "Estatus de factura", "VALOR DE FACTURA",
    "Estatus del pago de factura", "Fecha de solicitud de factura", "UUID",
    "Posting date",
]

# Columnas que deben quedar como número (el resto se trata como texto/fecha)
NUMERIC_COLUMNS = {
    "Year", "Month", "Número ICD contrato", "Solicitud de excepción",
    "VALOR DE FACTURA",
}

# Normalización de valores con typos/variantes que deben unificarse (case-insensitive)
AT_NORMALIZE = {
    "inmunology": "Inmunología",
    "inmunología": "Inmunología",
}


def normalize_at(v):
    if not isinstance(v, str):
        return None
    return AT_NORMALIZE.get(v.strip().lower(), v)


def clean(v):
    """Limpia celdas: convierte #REF!, cadenas vacías y espacios sobrantes en None."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s.upper() in ("#REF!", "#N/A", "#NAME?", "#VALUE!", "#DIV/0!"):
            return None
        return s
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()[:10]
    if isinstance(v, datetime.time):
        return v.isoformat()
    return v


def clean_number(v):
    v = clean(v)
    if v is None:
        return None
    try:
        n = float(v)
        # si es entero exacto, guardarlo como int para que se vea limpio en el JSON
        return int(n) if n.is_integer() else n
    except (TypeError, ValueError):
        return None


def pick_sheet(wb):
    # Ajusta este nombre si tu pestaña de eventos se llama distinto
    for name in ("Nacionales", "Eventos", "Sheet1"):
        if name in wb.sheetnames:
            return wb[name]
    return wb[wb.sheetnames[0]]


def load_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = pick_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [h.strip() if isinstance(h, str) else h for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    missing = [c for c in COLUMNS if c not in idx]
    if missing:
        print("Aviso: no se encontraron estas columnas en el Excel (se guardarán como null):")
        for m in missing:
            print(f"   - {m}")

    def get(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    records = []
    for row in rows[1:]:
        if not any(clean(v) is not None for v in row):
            continue  # fila totalmente vacía

        rec = {}
        for col in COLUMNS:
            raw = get(row, col)
            if col in NUMERIC_COLUMNS:
                rec[col] = clean_number(raw)
            elif col == "AT":
                rec[col] = normalize_at(clean(raw))
            else:
                rec[col] = clean(raw)
        records.append(rec)

    return records


def load_provisiones(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Provisiones" not in wb.sheetnames:
        return [], []
    ws = wb["Provisiones"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], []

    cols_nac = ["Estatus", "Cuenta", "MONTO MXN", "Centro de Costos", "Orden Interna",
                "Nombre del Evento", "Número del Evento", "Fecha de evento",
                "Fecha Firma de contrato", "No. de contrato / No. RN", "Concatenado",
                "AT", "Franquicia"]
    cols_inter = ["Estatus", "Cuenta", "MONTO", "Centro de Costos", "Orden Interna",
                  "Nombre del Evento", "Número del Evento", "Fecha de evento",
                  "Fecha Firma de contrato", "No. de contrato / No. RN", "Concatenado",
                  "AT", "Franquicia", "Divisa"]

    nac, inter = [], []
    for row in rows[1:]:
        block1 = row[0:13]
        block2 = row[15:29]

        if any(clean(v) is not None for v in block1):
            rec = {}
            for i, col in enumerate(cols_nac):
                raw = block1[i] if i < len(block1) else None
                if col == "MONTO MXN":
                    rec[col] = clean_number(raw)
                elif col == "AT":
                    rec[col] = normalize_at(clean(raw))
                else:
                    rec[col] = clean(raw)
            nac.append(rec)

        if any(clean(v) is not None for v in block2):
            rec = {}
            for i, col in enumerate(cols_inter):
                raw = block2[i] if i < len(block2) else None
                if col == "MONTO":
                    rec[col] = clean_number(raw)
                elif col == "AT":
                    rec[col] = normalize_at(clean(raw))
                else:
                    rec[col] = clean(raw)
            inter.append(rec)

    return nac, inter


def inject_data(template, marker, records):
    start_marker = template.find(marker)
    if start_marker == -1:
        print(f'Aviso: no encontré "{marker}" en el template, se omite.')
        return template

    json_start = start_marker + len(marker)
    decoder = json.JSONDecoder()
    try:
        _, json_end = decoder.raw_decode(template, json_start)
    except json.JSONDecodeError:
        print(f'Error: no pude leer el arreglo existente para "{marker}".')
        sys.exit(1)

    data_json = json.dumps(records, ensure_ascii=False)
    return template[:json_start] + data_json + template[json_end:]


def main():
    if len(sys.argv) < 2:
        print("Uso: python3 build_dashboard.py ruta/al/excel.xlsx [salida.html]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else "dashboard.html"

    records = load_rows(xlsx_path)
    prov_nac, prov_inter = load_provisiones(xlsx_path)

    with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
        template = f.read()

    template = inject_data(template, "const EVENTOS_RAW = ", records)
    template = inject_data(template, "const PROVISIONES_NAC_RAW = ", prov_nac)
    template = inject_data(template, "const PROVISIONES_INTER_RAW = ", prov_inter)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(template)

    print(f"OK: {len(records)} eventos, {len(prov_nac)} provisiones nacionales, "
          f"{len(prov_inter)} provisiones internacionales -> {out_path}")


if __name__ == "__main__":
    main()
